import sqlite3
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.db.session import rows_to_dicts


LEASE_FIELDS = {
    "city",
    "community_name",
    "address",
    "acreage",
    "price",
    "listing_date",
    "rental_type",
    "recorder",
    "maintainor",
    "has_key",
    "agent",
    "deal_date",
    "lease_expire_date",
    "cancel_time",
    "cancel_reason",
    "for_sale",
    "owner_phone",
    "customer_phone",
    "status",
    "creator",
    "modifier",
}

HEADER_MAP = {
    "小区名称": "community_name",
    "房源地址": "address",
    "面积": "acreage",
    "价格": "price",
    "挂牌时间": "listing_date",
    "出租方式": "rental_type",
    "录入人": "recorder",
    "维护人": "maintainor",
    "是否有钥匙": "has_key",
    "成交人": "agent",
    "成交日期": "deal_date",
    "租期到期时间": "lease_expire_date",
    "核销时间": "cancel_time",
    "核销原因": "cancel_reason",
    "是否出售": "for_sale",
    "业主电话": "owner_phone",
    "客户电话": "customer_phone",
}


def _date_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace("/", ".").replace("-", ".").strip(".")
    parts = [part for part in normalized.split(".") if part]
    if len(parts) >= 3 and all(part.isdigit() for part in parts[:3]):
        return f"{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}"
    if len(parts) == 2 and all(part.isdigit() for part in parts):
        return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
    return text


def _float_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def _bool_value(value: Any) -> int | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if text in {"是", "有", "true", "True", "1", "已出售"}:
        return 1
    if text in {"否", "无", "false", "False", "0", "未出售"}:
        return 0
    return None


def _text_value(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value).strip()


def normalize_lease_payload(data: dict[str, Any], city: str, username: str | None = None) -> dict[str, Any]:
    payload = {key: value for key, value in data.items() if key in LEASE_FIELDS}
    payload["city"] = payload.get("city") or city
    for key in ["acreage", "price"]:
        payload[key] = _float_value(payload.get(key))
    for key in ["has_key", "for_sale"]:
        payload[key] = _bool_value(payload.get(key))
    for key in ["listing_date", "deal_date", "lease_expire_date", "cancel_time"]:
        payload[key] = _date_text(payload.get(key))
    for key in [
        "community_name",
        "address",
        "rental_type",
        "recorder",
        "maintainor",
        "agent",
        "cancel_reason",
        "owner_phone",
        "customer_phone",
        "status",
    ]:
        payload[key] = _text_value(payload.get(key))
    if not payload.get("status"):
        payload["status"] = "active"
    if username:
        payload.setdefault("creator", username)
    return payload


def create_lease(conn: sqlite3.Connection, data: dict[str, Any], city: str, username: str) -> int:
    payload = normalize_lease_payload(data, city, username)
    columns = ", ".join(payload.keys())
    placeholders = ", ".join(["?"] * len(payload))
    cursor = conn.execute(f"INSERT INTO lease_properties({columns}) VALUES ({placeholders})", tuple(payload.values()))
    return int(cursor.lastrowid)


def update_lease(conn: sqlite3.Connection, lease_id: int, data: dict[str, Any], city: str, username: str) -> dict | None:
    existing = get_lease(conn, lease_id, city)
    if not existing:
        return None
    payload = normalize_lease_payload(data, city)
    payload["modifier"] = username
    payload["modify_time"] = datetime.now().isoformat(timespec="seconds")
    assignments = ", ".join([f"{key} = ?" for key in payload])
    conn.execute(f"UPDATE lease_properties SET {assignments} WHERE id = ? AND city = ?", tuple(payload.values()) + (lease_id, city))
    return get_lease(conn, lease_id, city)


def delete_lease(conn: sqlite3.Connection, lease_id: int, city: str, username: str) -> bool:
    cursor = conn.execute(
        """
        UPDATE lease_properties
        SET status = 'deleted', modifier = ?, modify_time = CURRENT_TIMESTAMP
        WHERE id = ? AND city = ? AND status != 'deleted'
        """,
        (username, lease_id, city),
    )
    return cursor.rowcount > 0


def get_lease(conn: sqlite3.Connection, lease_id: int, city: str) -> dict | None:
    row = conn.execute(
        "SELECT * FROM lease_properties WHERE id = ? AND city = ? AND status != 'deleted'",
        (lease_id, city),
    ).fetchone()
    return dict(row) if row else None


def list_leases(
    conn: sqlite3.Connection,
    city: str,
    community_name: str | None = None,
    price_min: float | None = None,
    price_max: float | None = None,
    sort_by: str = "lease_expire_date",
    sort_order: str = "asc",
    page: int = 1,
    page_size: int = 20,
) -> dict[str, Any]:
    clauses = ["city = ?", "status != 'deleted'"]
    params: list[Any] = [city]
    if community_name:
        clauses.append("community_name LIKE ?")
        params.append(f"%{community_name}%")
    if price_min is not None:
        clauses.append("price >= ?")
        params.append(price_min)
    if price_max is not None:
        clauses.append("price <= ?")
        params.append(price_max)

    where_sql = " AND ".join(clauses)
    safe_sort_by = sort_by if sort_by in {"price", "lease_expire_date"} else "lease_expire_date"
    safe_sort_order = "DESC" if sort_order.lower() == "desc" else "ASC"
    safe_page = max(page, 1)
    safe_page_size = min(max(page_size, 10), 200)
    offset = (safe_page - 1) * safe_page_size

    total = conn.execute(f"SELECT COUNT(*) AS total FROM lease_properties WHERE {where_sql}", tuple(params)).fetchone()["total"]
    rows = conn.execute(
        f"""
        SELECT *
        FROM lease_properties
        WHERE {where_sql}
        ORDER BY {safe_sort_by} {safe_sort_order}, id DESC
        LIMIT ? OFFSET ?
        """,
        tuple(params + [safe_page_size, offset]),
    ).fetchall()
    return {"items": rows_to_dicts(rows), "total": total, "page": safe_page, "page_size": safe_page_size}


def import_leases_from_excel(conn: sqlite3.Connection, content: bytes, city: str, username: str) -> dict:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index_map = {HEADER_MAP[header]: index for index, header in enumerate(headers) if header in HEADER_MAP}
    if "community_name" not in index_map:
        raise ValueError("Excel 缺少小区名称列")

    total = success = skipped = failed = 0
    errors = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        total += 1
        try:
            data = {field: row[index] if index < len(row) else None for field, index in index_map.items()}
            payload = normalize_lease_payload(data, city, username)
            if not payload.get("community_name"):
                skipped += 1
                continue
            duplicate = conn.execute(
                """
                SELECT id
                FROM lease_properties
                WHERE city = ?
                  AND community_name = ?
                  AND COALESCE(address, '') = COALESCE(?, '')
                  AND COALESCE(deal_date, '') = COALESCE(?, '')
                  AND status != 'deleted'
                LIMIT 1
                """,
                (city, payload.get("community_name"), payload.get("address"), payload.get("deal_date")),
            ).fetchone()
            if duplicate:
                skipped += 1
                continue
            columns = ", ".join(payload.keys())
            placeholders = ", ".join(["?"] * len(payload))
            conn.execute(f"INSERT INTO lease_properties({columns}) VALUES ({placeholders})", tuple(payload.values()))
            success += 1
        except Exception as exc:
            failed += 1
            errors.append({"row": row_number, "error": str(exc)})
    conn.commit()
    return {"total": total, "success": success, "skipped": skipped, "failed": failed, "errors": errors[:20]}
