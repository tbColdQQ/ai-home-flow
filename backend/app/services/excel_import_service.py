import json
import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.services.order_service import create_order


HEADER_MAP = {
    "城市": "city",
    "区域": "area",
    "街道": "street",
    "小区名称": "residential",
    "小区": "residential",
    "楼盘": "residential",
    "面积(㎡)": "acreage",
    "面积": "acreage",
    "成交价(万)": "price",
    "成交价": "price",
    "成交价格": "price",
    "成交人": "agent",
    "经纪人": "agent",
    "门店": "store",
    "品牌": "brand",
    "签约时间": "signing_date",
    "签约日期": "signing_date",
    "成交日期": "signing_date",
    "维护人": "maintainor",
    "维护人CA": "CA",
    "CA": "CA",
    "车位": "parking",
    "备注": "remark",
}


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().replace(" ", "")


def _parse_price(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value) * 10000
    text = str(value).strip().replace(",", "")
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group(0))
    return number * 10000 if "万" in text else number


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group(0)) if match else None


def _parse_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    match = re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", text)
    if match:
        return _parse_date(match.group(0))
    return None


def _parse_parking(value: Any) -> int | None:
    if value is None or value == "":
        return None
    text = str(value).strip()
    if text in {"有", "是", "1", "true", "True"}:
        return 1
    if text in {"无", "否", "0", "false", "False"}:
        return 0
    return None


def _has_duplicate_order(conn, data: dict[str, Any]) -> bool:
    row = conn.execute(
        """
        SELECT ID
        FROM orders
        WHERE city = ?
          AND signing_date = ?
          AND residential = ?
          AND ABS(COALESCE(acreage, 0) - ?) < 0.001
          AND ABS(COALESCE(price, 0) - ?) < 0.001
          AND COALESCE(status, 'normal') = 'normal'
        LIMIT 1
        """,
        (
            data.get("city"),
            data.get("signing_date"),
            data.get("residential"),
            data.get("acreage") or 0,
            data.get("price") or 0,
        ),
    ).fetchone()
    return row is not None


def import_orders_from_excel(conn, file_bytes: bytes, file_name: str, operator: str | None = None) -> dict:
    cursor = conn.execute(
        """
        INSERT INTO import_batches(import_type, file_path, status)
        VALUES ('excel_orders', ?, 'running')
        """,
        (file_name,),
    )
    batch_id = int(cursor.lastrowid)
    total = success = skipped = failed = 0
    errors: list[dict[str, Any]] = []

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValueError("Excel 没有表头")

        field_indexes: dict[int, str] = {}
        for index, header in enumerate(headers):
            field = HEADER_MAP.get(_normalize_header(header))
            if field:
                field_indexes[index] = field

        for row_number, row in enumerate(rows, start=2):
            if not any(cell not in (None, "") for cell in row):
                continue
            total += 1
            raw: dict[str, Any] = {}
            for index, field in field_indexes.items():
                raw[field] = row[index] if index < len(row) else None

            data = {
                "city": str(raw.get("city") or "").strip() or None,
                "area": str(raw.get("area") or "").strip() or None,
                "street": str(raw.get("street") or "").strip() or None,
                "residential": str(raw.get("residential") or "").strip() or None,
                "acreage": _parse_float(raw.get("acreage")),
                "price": _parse_price(raw.get("price")),
                "agent": str(raw.get("agent") or "").strip() or None,
                "store": str(raw.get("store") or "").strip() or None,
                "brand": str(raw.get("brand") or "").strip() or None,
                "signing_date": _parse_date(raw.get("signing_date")),
                "maintainor": str(raw.get("maintainor") or "").strip() or None,
                "CA": str(raw.get("CA") or "").strip() or None,
                "parking": _parse_parking(raw.get("parking")),
                "remark": str(raw.get("remark") or "").strip() or None,
                "status": "normal",
                "source_type": "excel",
                "source_id": batch_id,
                "source_file": file_name,
                "review_status": "confirmed",
            }
            data["raw_payload_json"] = json.dumps(raw, ensure_ascii=False, default=str)

            missing = [field for field in ["city", "residential", "acreage", "price", "signing_date"] if not data.get(field)]
            if missing:
                failed += 1
                errors.append({"row": row_number, "reason": f"必填字段缺失：{','.join(missing)}"})
                continue

            if _has_duplicate_order(conn, data):
                skipped += 1
                continue

            data["creator"] = operator
            create_order(conn, data)
            success += 1

        conn.execute(
            """
            UPDATE import_batches
            SET status = 'done',
                total_count = ?,
                success_count = ?,
                failed_count = ?,
                error_json = ?,
                finish_time = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (total, success, failed, json.dumps({"skipped": skipped, "errors": errors[:50]}, ensure_ascii=False), batch_id),
        )
        conn.commit()
        return {"batch_id": batch_id, "total": total, "success": success, "failed": failed, "skipped": skipped, "errors": errors[:20]}
    except Exception as exc:
        conn.execute(
            """
            UPDATE import_batches
            SET status = 'failed', error_json = ?, finish_time = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (json.dumps({"error": str(exc)}, ensure_ascii=False), batch_id),
        )
        conn.commit()
        raise
